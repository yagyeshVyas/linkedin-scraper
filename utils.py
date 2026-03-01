"""
Utility Functions for LinkedIn Scraper
=======================================
Helper functions for delays, scrolling, file I/O, and Excel export.
"""

import asyncio
import json
import logging
import random
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────
#  HUMAN-LIKE BEHAVIOR
# ─────────────────────────────────────────────

async def human_delay(min_sec: float = 1.0, max_sec: float = 3.0):
    """Sleep for a random duration to mimic human behavior."""
    delay = random.uniform(min_sec, max_sec)
    await asyncio.sleep(delay)


async def random_scroll(page, scrolls: int = 4):
    """Scroll the page randomly like a human reading content."""
    for _ in range(scrolls):
        scroll_amount = random.randint(300, 800)
        direction = random.choice([1, 1, 1, -1])  # mostly down
        await page.evaluate(f"window.scrollBy(0, {scroll_amount * direction})")
        await asyncio.sleep(random.uniform(0.5, 1.5))


# ─────────────────────────────────────────────
#  PROGRESS SAVE/LOAD (Resume after crash)
# ─────────────────────────────────────────────

def save_progress(filepath: str, data: dict):
    """Save scraping progress to JSON for resume capability."""
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    logger.debug(f"Progress saved: {len(data.get('results', []))} records")


def load_progress(filepath: str) -> dict:
    """Load previous scraping progress if it exists."""
    if Path(filepath).exists():
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        logger.info(f"📂 Resumed progress: {len(data.get('results', []))} existing records, "
                    f"{len(data.get('completed_companies', []))} companies done")
        return data
    return {"results": [], "completed_companies": []}


# ─────────────────────────────────────────────
#  LOGGING SETUP
# ─────────────────────────────────────────────

def setup_logging(log_file: str = "scraper.log"):
    """Configure logging to both console and file."""
    Path(log_file).parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(log_file, encoding="utf-8")
        ]
    )


# ─────────────────────────────────────────────
#  EXCEL EXPORT (Beautiful formatting)
# ─────────────────────────────────────────────

def export_to_excel(results: list, output_path: str) -> str:
    """
    Export scraped recruiter data to a beautifully formatted Excel file.
    Includes clickable LinkedIn URLs, color-coded rows, and auto-sized columns.
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Build DataFrame
    df = pd.DataFrame(results)

    # Reorder and rename columns
    columns_order = [
        "name", "title", "company", "search_company",
        "headline", "location", "email", "phone",
        "connections", "linkedin_url", "scraped_at"
    ]
    # Keep only existing columns
    columns_order = [c for c in columns_order if c in df.columns]
    df = df[columns_order]

    column_names = {
        "name": "Full Name",
        "title": "Job Title",
        "company": "Company (from Profile)",
        "search_company": "Searched Company",
        "headline": "LinkedIn Headline",
        "location": "Location",
        "email": "Email Address",
        "phone": "Phone",
        "connections": "Connections",
        "linkedin_url": "LinkedIn URL",
        "scraped_at": "Scraped At"
    }
    df = df.rename(columns=column_names)

    # Write to Excel
    df.to_excel(output_path, index=False, sheet_name="Recruiters")

    # ── Apply Beautiful Formatting ──
    wb = load_workbook(output_path)
    ws = wb.active

    # Colors
    HEADER_BG = "0A66C2"     # LinkedIn blue
    HEADER_FG = "FFFFFF"     # White
    ROW_ALT   = "EBF3FB"     # Light blue
    ROW_NORM  = "FFFFFF"     # White
    BORDER_CLR = "CCCCCC"

    thin_border = Border(
        left=Side(style="thin", color=BORDER_CLR),
        right=Side(style="thin", color=BORDER_CLR),
        top=Side(style="thin", color=BORDER_CLR),
        bottom=Side(style="thin", color=BORDER_CLR)
    )

    # Style header row
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 32

    # Style data rows + make URLs clickable
    url_col_idx = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "LinkedIn URL":
            url_col_idx = col_idx
            break

    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
        fill_color = ROW_ALT if row_num % 2 == 0 else ROW_NORM
        for col_idx, cell in enumerate(row, 1):
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = thin_border
            cell.font = Font(size=10, name="Calibri")

            # Make LinkedIn URL clickable hyperlink
            if url_col_idx and col_idx == url_col_idx and cell.value:
                url = str(cell.value)
                cell.hyperlink = url
                cell.value = url
                cell.font = Font(
                    size=10, color="0563C1", underline="single",
                    name="Calibri"
                )

        ws.row_dimensions[row_num].height = 22

    # Auto-size columns
    col_widths = {
        "Full Name": 22,
        "Job Title": 28,
        "Company (from Profile)": 25,
        "Searched Company": 22,
        "LinkedIn Headline": 40,
        "Location": 22,
        "Email Address": 28,
        "Phone": 16,
        "Connections": 14,
        "LinkedIn URL": 45,
        "Scraped At": 18,
    }
    for col_idx, cell in enumerate(ws[1], 1):
        col_letter = get_column_letter(col_idx)
        width = col_widths.get(cell.value, 18)
        ws.column_dimensions[col_letter].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "LinkedIn Recruiter Scrape - Summary Report"
    ws_summary["A1"].font = Font(bold=True, size=14, color="0A66C2")
    ws_summary["A3"] = "Generated On:"
    ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A4"] = "Total Recruiters Found:"
    ws_summary["B4"] = len(results)
    ws_summary["A5"] = "Unique Companies:"
    ws_summary["B5"] = len(set(r.get("search_company", "") for r in results))
    ws_summary["A6"] = "Profiles with Email:"
    ws_summary["B6"] = sum(1 for r in results if r.get("email"))

    for row in ws_summary.iter_rows(min_row=3, max_row=6, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(size=11, name="Calibri")
            if cell.column == 1:
                cell.font = Font(bold=True, size=11, name="Calibri")
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 22

    wb.save(output_path)
    logger.info(f"✅ Excel file saved: {output_path}")
    return output_path